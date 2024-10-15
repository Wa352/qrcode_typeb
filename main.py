from flask import Flask, request, render_template, make_response
from pyzbar.pyzbar import decode
from PIL import Image
import os
import json
import openpyxl as op



app = Flask(__name__)

def cell_find(path, info):  # 項目のセルを見つけるプログラム
    wb = op.load_workbook(path)  # Excelファイルを開く
    ws = wb.active  # アクティブなワークシートを取得
    hight_counter = 1  # 行カウンターの初期化
    while True:
        hight_counter += 1   # 登録されているセルを探す
        hight_outputer = str(hight_counter)
        value = ws["A" + hight_outputer].value  # 列Aの値を取得
        if value == None:   # 登録されていなかった場合
            flag = False  # フラグをFalseに設定
            score = None
            break
        if value == info:  # 登録されている場合
            flag = True  # フラグをTrueに設定
            score = int(ws["B" + hight_outputer].value)  # 列Bのスコアを取得
            break
    wb.close()  # ファイルを閉じる
    return hight_counter, flag, score  # 行カウンターとフラグを返す

def read_qr_code(image_path):
    image = Image.open(image_path)
    decoded_objects = decode(image)
    os.remove(image_path)
    if decoded_objects:
        return decoded_objects[0].data.decode('utf-8')
    return None

UPLOAD_FOLDER = './uploads'
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

@app.route('/', methods=['GET', 'POST'])
def index():
    path = "./score.xlsx"
    if request.method == 'POST':
        file = request.files.get('file')
        if file:
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
            file.save(file_path)
            info = read_qr_code(file_path)
            if info:
                hight_counter,flag,score = cell_find(path, info)
                print(score)
                print(flag)
                if flag:
                    responce = make_response(render_template("select.html", score=score))
                    responce.set_cookie("id", value=json.dumps({"info":info, "conter":hight_counter}))
                    return responce
                else:
                    responce = make_response(render_template("add.html"))
                    responce.set_cookie("id", value=json.dumps({"info":info, "conter":hight_counter}))
                    return responce
            else:
                return "qrcode not found"
            
        # ファイルがアップロードされなかった場合
        return "ファイルが選択されていません。", 400
    
    # GETリクエストの場合はフォームを表示
    return render_template("camera.html")

@app.route('/add')
def add():
    cookie = request.cookies.get("id")
    if cookie:
        cookie = json.loads(cookie)
        info = cookie["info"]
        hight_counter = cookie["conter"]
        path = "./score.xlsx"
        wb = op.load_workbook(path)  # Excelファイルを開く
        ws = wb.active  # アクティブなワークシートを取得
        hight_outputer = str(hight_counter)
        ws["A" + hight_outputer] = info  # 情報を列Aに入力
        ws["B" + hight_outputer] = 0  # 初期スコアを0に設定
        wb.save(path)  # ファイルを保存
        return render_template("add_success.html")
    else:
        return "cookie not found"


@app.route('/change', methods=['GET', 'POST'])
def change():
    path = "./score.xlsx"
    if request.method == 'POST':
        cookie = request.cookies.get("id")
        new_score = request.form.get('score')
        print(new_score)
        if cookie:
            cookie = json.loads(cookie)
            hight_counter = cookie["conter"]
            wb = op.load_workbook(path)  # Excelファイルを開く
            ws = wb.active  # アクティブなワークシートを取得
            hight_outputer = str(hight_counter)
            ws["B" + hight_outputer] = int(new_score)  # 新しいスコアを入力
            wb.save(path)  # ファイルを保存
            return render_template("change_success.html")
    return render_template("change.html")

@app.route('/delete')
def delete():
    cookie = request.cookies.get("id")

    path = "./score.xlsx"
    cookie = json.loads(cookie)
    hight_counter = cookie["conter"]
    hight_outputer = str(hight_counter)
    wb = op.load_workbook(path)  # Excelファイルを開く
    ws = wb.active  # アクティブなワークシートを取得
    hight_outputer_start = hight_outputer
    hight_counter_stop = hight_counter
    ws["A" + hight_outputer] = None  # 列Aの値を削除
    ws["B" + hight_outputer] = None  # 列Bの値を削除
    while True:  # 削除した後、セルを詰める
        hight_counter_stop += 1
        hight_outputer_stop = str(hight_counter_stop)
        value = ws["A" + hight_outputer_stop].value   
        if value == None:
            # 範囲を移動して詰める
            ws.move_range('A' + hight_outputer_start + ":" + "B" + hight_outputer_stop, rows=1, cols=0, translate=True) 
            break
    wb.save(path)  # ファイルを保存
    return render_template("delete.html")


if __name__ == "__main__":
    app.run(debug=True)
